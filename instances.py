from openpyxl import Workbook
from tkinter.filedialog import askdirectory
from openpyxl.styles import Font, Alignment, PatternFill

class AppCore:
    __DriverSchedule = 'DriverSchedule'
    __SingIn = 'SingIn' 
    __Main = 'Main'

    def getDriverSchedule(self):
        return self.__DriverSchedule
    
    def getSingIn(self):
        return self.__SingIn
    
    def getMain(self):
        return self.__Main

    __registered_input = {
        __DriverSchedule: [],
        __SingIn:         []
    }
    __registered_windows  = {
        __DriverSchedule: [],
        __SingIn:         [],
        __Main:           []
    }
    __registered_winget = {
        __DriverSchedule: [],
        __SingIn:         [],
        __Main:           []
    }

    def getRegisteredInput(self):
        return self.__registered_input
    
    def getRegisteredWindow(self):
        return self.__registered_windows
    
    def getRegisteredWinget(self):
        return self.__registered_winget
    
    def setRegisteredInput(self, id, input):
        self.__registered_input[id].append(input)

    def setRegisteredWindow(self, id, window):
        self.__registered_windows[id].append(window)

    def setRegisteredWinget(self, id, winget):
        self.__registered_winget[id].append(winget)
    
    def unsubscribeInput(self, id):
        for key in self.__registered_input[id]:
            key = None
            self.__registered_input[id].clear()

    def unsubscribeWindow(self, id):
        for key in self.__registered_windows[id]:
            key = None
            self.__registered_windows[id].clear()
    
    def unsubscribeWinget(self, id):
        for key in self.__registered_winget[id]:
            key = None
            self.__registered_winget[id].clear()


class MotoristasDados:
    __id_name = "NOME MOTORISTA"
    __id_start_load = "CHEGADA CARREGAMENTO"
    __id_end_load = "SAIDA CARREGAMENTO"
    __id_start_unload = "CHEGADA DESCARGA"
    __id_end_unload = "SAIDA DESCARGA"


    __data_state = {
            __id_name:         [],
            __id_start_load:   [],
            __id_end_load:     [],
            __id_start_unload: [],
            __id_end_unload:   []
        }
    
    def get_data_state(self):
        return self.__data_state
    
    def get_id_state(self):
        __listKeys = []
        __listState = self.get_data_state()
        for key in __listState.keys():
            __listKeys.append(key)

        return __listKeys

    # Acima metodos Get n' Set

    def clearDataBase(self):
            dataState = self.get_data_state()
            idName = self.get_id_state()
            cd1 = list(dataState[idName[0]])
            cd2 = list(dataState[idName[1]])
            cd3 = list(dataState[idName[2]])
            cd4 = list(dataState[idName[3]])
            cd5 = list(dataState[idName[4]])

            if len(dataState[idName[0]]) != 0:
                for self.numberFrame in cd1:
                    dataState[idName[0]].pop()

            if len(dataState[idName[1]]) != 0:
                for self.numberFrame in cd2:
                    dataState[idName[1]].pop()

            if len(dataState[idName[2]]) != 0:
                for self.numberFrame in cd3:
                    dataState[idName[2]].pop()

            if len(dataState[idName[3]]) != 0:
                for self.numberFrame in cd4:
                    dataState[idName[3]].pop()

            if len(dataState[idName[4]]) != 0:
                for self.numberFrame in cd5:
                    dataState[idName[4]].pop()

    def saveSheetInDirectory(self,file):
        core = AppCore()
        registerWinget = core.getRegisteredWinget()
        key = core.getDriverSchedule()
        directorySave = askdirectory(title='Selecione onde deseja salvar o arquivo')

        if directorySave != '':
            file.save(directorySave + '/Horarios Motoristas.xlsx')
            registerWinget[key][6]['text'] = 'Planilha Gerada!'

        else:
            registerWinget[key][6]['text'] = 'Selecione uma pasta para salvar o arquivo'

    def creatSheet(self):
        core = AppCore()
        registerWinget = core.getRegisteredWinget()
        key = core.getDriverSchedule()
        dataState = self.get_data_state()
        dataName = self.get_id_state()
        cs1 = len(dataState[dataName[0]])
        cs2 = len(dataState[dataName[1]])
        cs3 = len(dataState[dataName[2]])
        cs4 = len(dataState[dataName[3]])
        cs5 = len(dataState[dataName[4]])
 
        if cs1 and cs2 and cs3 and cs4 and cs5 != 0:
            file = Workbook()
            sheet = file.active
            sheet.title = 'OP TRANSPORTE'

            names_column = list(dataState.keys())
            print(names_column)

            for number_column, column in enumerate(names_column, start = 1):
                cell = sheet.cell(row = 1, column = number_column)
                cell.value = column

                cell.fill = PatternFill(start_color = '0000CCFF', end_color = '0000CCFF', fill_type = 'solid')
                cell.font = Font(bold = True, size = 13, name = 'Arial', color = '000000')
                cell.alignment = Alignment(horizontal = 'center')

                value_column = dataState[column]

                for number_value, value in enumerate(value_column, start = 2):
                    cell = sheet.cell(row = number_value, column = number_column)
                    cell.value = value

                    cell.font = Font(name = 'Arial', size = 12)
                    cell.alignment = Alignment(horizontal = 'center')

            self.saveSheetInDirectory(file)
            self.clearDataBase()
        else:
            registerWinget[key][6]['text'] = 'Insira todas as informações!'

    def insert(self):
        core = AppCore()
        dataState = self.get_data_state()
        dataName = self.get_id_state()
        key = core.getDriverSchedule()
        self.text_insert = 'Insira as Informações'

        listInput = core.getRegisteredInput()
        registerWinget = core.getRegisteredWinget()

        if listInput[key][0].get() != '':
            dataState[dataName[0]].append(listInput[key][0].get())
            listInput[key][0].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][1].get() != '':
            dataState[dataName[1]].append(listInput[key][1].get())
            listInput[key][1].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][2].get() != '':
            dataState[dataName[2]].append(listInput[key][2].get())
            listInput[key][2].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][3].get() != '':
            dataState[dataName[3]].append(listInput[key][3].get())
            listInput[key][3].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][4].get() != '':
            dataState[dataName[4]].append(listInput[key][4].get())
            listInput[key][4].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()


class PessoaDados:
    processList = []

    __id_cadastro_nome = "NOME"
    __id_cadastro_sobrenome = "SOBRENOME"
    __id_cadastro_data_nascimento = "DATA NASCIMENTO"
    __id_cadastro_numero_telefone = "NUMERO TELEFONE" 
    __id_cadastro_email = "E-MAIL"
    __id_cadastro_logradouro = "LOGRADOURO"
    __id_cadastro_numero_casa = "NÚMERO DA CASA"
    __id_cadastro_complemento = "COMPLEMENTO"
    __id_cadastro_bairro = "BAIRRO" 
    __id_cadastro_cidade = "CIDADE"
    __id_cadastro_cep = "CEP"
    __id_cadastro_estado = "ESTADO"

    __data_for_signup = {
        __id_cadastro_nome:            [],
        __id_cadastro_sobrenome:       [],
        __id_cadastro_data_nascimento: [],
        __id_cadastro_numero_telefone: [],
        __id_cadastro_email:           [],
        __id_cadastro_logradouro:      [],
        __id_cadastro_numero_casa:     [],
        __id_cadastro_complemento:     [],
        __id_cadastro_bairro:          [],
        __id_cadastro_cidade:          [],
        __id_cadastro_cep:             [],
        __id_cadastro_estado:          []
    }

    def get_data_for_singup(self):
        return self.__data_for_signup
    
    def get_id_cadastros(self):
        __id_cadastros = []
        __dataSignup = self.get_data_for_singup()

        for i in __dataSignup.keys():
            __id_cadastros.append(i)

        return __id_cadastros
    
    def set_data_for_singup(self, name, key, num):
        core = AppCore()
        registerInput = core.getRegisteredInput()

        listDataUser = self.__data_for_signup

        listDataUser[name].append(registerInput[key][num].get())

    # Acima metodos Get n' Set
    
    def insert(self):
        core = AppCore()
        registerInput = core.getRegisteredInput()
        key = core.getSingIn()
        __listDataUser = self.get_data_for_singup()

        for id in __listDataUser:
            self.processList.append(id)
        
        for i, id in enumerate(registerInput[key]):
            if registerInput[key][i].get() != '':
                self.set_data_for_singup(self.processList[i], key, i)
                print(f'{__listDataUser[self.processList[i]][0]} --> Pré Carregado\n')
                #print(self.processList[i], key)
            else:
                print('PRENENCHA OS CAMPOS!')


    def clearData(self):
        core = AppCore()
        registerInput = core.getRegisteredInput()
        key = core.getSingIn()
        __listDataUser = self.get_data_for_singup()

        for id in __listDataUser:
            for num, x in enumerate(__listDataUser[id]): # retirar laço for
                x == None
                __listDataUser[id].pop(0)

        self.processList.clear()
        
        for num, x in enumerate(registerInput[key]):
            registerInput[key][num].delete(0, 'end')
        
        print('INSIRA TODAS AS INSFORMAÇÕES\nINTERFACE LIMPA\n')

class ConsoleRedirect:
    core = AppCore()
    windowR = core.getRegisteredWindow()
    key = core.getMain()

    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.insert('end', message)   
        self.text_widget.see('end')  

    def flush(self):
        pass
