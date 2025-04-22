from openpyxl import Workbook
from tkinter.filedialog import askdirectory
from openpyxl.styles import Font, Alignment, PatternFill
from logic.AppCore import AppCore

class DataDriver:
    __id_name = "NOME MOTORISTA"
    __id_start_load = "CHEGADA CARREGAMENTO"
    __id_end_load = "SAIDA CARREGAMENTO"
    __id_start_unload = "CHEGADA DESCARGA"
    __id_end_unload = "SAIDA DESCARGA"


    __data_state = {
            __id_name:         [],
            __id_start_load:   [],
            __id_end_load:     [],
            __id_start_unload: [],
            __id_end_unload:   []
        }
    
    def get_data_state(self):
        return self.__data_state
    
    def get_id_state(self):
        __listKeys = []
        __listState = self.get_data_state()
        for key in __listState.keys():
            __listKeys.append(key)

        return __listKeys

    # Acima metodos Get n' Set

    def clearDataBase(self):
            dataState = self.get_data_state()
            idName = self.get_id_state()
            cd1 = list(dataState[idName[0]])
            cd2 = list(dataState[idName[1]])
            cd3 = list(dataState[idName[2]])
            cd4 = list(dataState[idName[3]])
            cd5 = list(dataState[idName[4]])

            if len(dataState[idName[0]]) != 0:
                for self.numberFrame in cd1:
                    dataState[idName[0]].pop()

            if len(dataState[idName[1]]) != 0:
                for self.numberFrame in cd2:
                    dataState[idName[1]].pop()

            if len(dataState[idName[2]]) != 0:
                for self.numberFrame in cd3:
                    dataState[idName[2]].pop()

            if len(dataState[idName[3]]) != 0:
                for self.numberFrame in cd4:
                    dataState[idName[3]].pop()

            if len(dataState[idName[4]]) != 0:
                for self.numberFrame in cd5:
                    dataState[idName[4]].pop()

    def saveSheetInDirectory(self,file):
        core = AppCore()
        registerWinget = core.getRegisteredWinget()
        key = core.getDriverSchedule()
        directorySave = askdirectory(title='Selecione onde deseja salvar o arquivo')

        if directorySave != '':
            file.save(directorySave + '/Horarios Motoristas.xlsx')
            registerWinget[key][6]['text'] = 'Planilha Gerada!'

        else:
            registerWinget[key][6]['text'] = 'Selecione uma pasta para salvar o arquivo'

    def creatSheet(self):
        core = AppCore()
        registerWinget = core.getRegisteredWinget()
        key = core.getDriverSchedule()
        dataState = self.get_data_state()
        dataName = self.get_id_state()
        cs1 = len(dataState[dataName[0]])
        cs2 = len(dataState[dataName[1]])
        cs3 = len(dataState[dataName[2]])
        cs4 = len(dataState[dataName[3]])
        cs5 = len(dataState[dataName[4]])
 
        if cs1 and cs2 and cs3 and cs4 and cs5 != 0:
            file = Workbook()
            sheet = file.active
            sheet.title = 'OP TRANSPORTE'

            names_column = list(dataState.keys())
            print(names_column)

            for number_column, column in enumerate(names_column, start = 1):
                cell = sheet.cell(row = 1, column = number_column)
                cell.value = column

                cell.fill = PatternFill(start_color = '0000CCFF', end_color = '0000CCFF', fill_type = 'solid')
                cell.font = Font(bold = True, size = 13, name = 'Arial', color = '000000')
                cell.alignment = Alignment(horizontal = 'center')

                value_column = dataState[column]

                for number_value, value in enumerate(value_column, start = 2):
                    cell = sheet.cell(row = number_value, column = number_column)
                    cell.value = value

                    cell.font = Font(name = 'Arial', size = 12)
                    cell.alignment = Alignment(horizontal = 'center')

            self.saveSheetInDirectory(file)
            self.clearDataBase()
        else:
            registerWinget[key][6]['text'] = 'Insira todas as informações!'

    def insert(self):
        core = AppCore()
        dataState = self.get_data_state()
        dataName = self.get_id_state()
        key = core.getDriverSchedule()
        self.text_insert = 'Insira as Informações'

        listInput = core.getRegisteredInput()
        registerWinget = core.getRegisteredWinget()

        if listInput[key][0].get() != '':
            dataState[dataName[0]].append(listInput[key][0].get())
            listInput[key][0].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][1].get() != '':
            dataState[dataName[1]].append(listInput[key][1].get())
            listInput[key][1].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][2].get() != '':
            dataState[dataName[2]].append(listInput[key][2].get())
            listInput[key][2].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][3].get() != '':
            dataState[dataName[3]].append(listInput[key][3].get())
            listInput[key][3].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()

        if listInput[key][4].get() != '':
            dataState[dataName[4]].append(listInput[key][4].get())
            listInput[key][4].delete(0, 'end')
        else: 
            registerWinget[key][6]['text'] = 'Insira todas as informações!'
            self.clearDataBase()