from openpyxl import Workbook
from tkinter.filedialog import askdirectory
from openpyxl.styles import Font, Alignment, PatternFill
from instasces import *
from tkinter import *

class horario_motorista:
        def __init__(self):

            self.window = Tk()
            self.window.title('HR - Motorista') 
            self.window.iconbitmap(iconImage)
            self.window.resizable(False, False)

            self.word_init = Label(self.window, text= 'Insira os dados da operação abaixo', width= 40, bg= backgroundcolorFont)
            self.word_init['font'] = fontText
            self.word_init.grid(column = 0, columnspan= 1, row = 0, padx= padX, pady= padY)

            self.inputLabel = Label(self.window, text= id_name, bg= backgroundcolorFont)
            self.inputLabel['font'] = fontText
            self.inputLabel.grid(column= 0, row= 3)

            self.inputValue = Entry(self.window, bd= border, bg= backgroundcolorWidget, width= entryWidth)
            self.inputValue.grid(column= 0, row= 4)

            self.inputLabel2 = Label(self.window, text= id_start_load,  bg= backgroundcolorFont)
            self.inputLabel2['font'] = fontText
            self.inputLabel2.grid(column= 0, row= 5)

            self.inputValue2 = Entry(self.window, bd= border, bg= backgroundcolorWidget, width= entryWidth)
            self.inputValue2.grid(column= 0, row= 6)

            self.inputLabel3 = Label(self.window, text= id_end_load, bg= backgroundcolorFont)
            self.inputLabel3['font'] = fontText
            self.inputLabel3.grid(column= 0, row= 7)

            self.inputValue3 = Entry(self.window, bd= border, bg= backgroundcolorWidget, width= entryWidth)
            self.inputValue3.grid(column= 0, row= 8)

            self.inputLabel4 = Label(self.window, text= id_start_unload, bg= backgroundcolorFont)
            self.inputLabel4['font'] = fontText
            self.inputLabel4.grid(column= 0, row= 9)

            self.inputValue4 = Entry(self.window, bd= border, bg= backgroundcolorWidget,  width= entryWidth)
            self.inputValue4.grid(column= 0, row= 10)

            self.inputLabel5 = Label(self.window, text= id_end_unload, bg= backgroundcolorFont)
            self.inputLabel5['font'] = fontText
            self.inputLabel5.grid(column= 0, row= 11)

            self.inputValue5 = Entry(self.window, bd= border, bg= backgroundcolorWidget,  width= entryWidth)
            self.inputValue5.grid(column= 0, row= 12)

            self.button1 = Button(self.window, text='Inserir Dados', bd= border, width= buttonsWidth, command= self.insert)
            self.button1.grid(column= 0, row= 13, pady= padY)
            self.button1['bg'] = backgroundcolorWidget
            self.button1['font'] = fontText

            self.button2 = Button(self.window, text='Gerar Planilha', bd= border, width= buttonsWidth, command= self.creatSheet)
            self.button2.grid(column= 0, row= 14, pady= padY)
            self.button2['bg'] = backgroundcolorWidget
            self.button2['font'] = fontText

            self.end_word = Label(self.window, text='', bg= backgroundcolorFont)
            self.end_word['font'] = fontText
            self.end_word.grid(column= 0, row= 15, pady= padY)

            self.window.mainloop()

        def clearDataBase(self):
            cd1 = list(data[id_name])
            cd2 = list(data[id_start_load])
            cd3 = list(data[id_end_load])
            cd4 = list(data[id_start_unload])
            cd5 = list(data[id_end_unload])

            if len(data[id_name]) != 0:
                for self.numberFrame in cd1:
                    data[id_name].pop()

            if len(data[id_start_load]) != 0:
                for self.numberFrame in cd2:
                    data[id_start_load].pop()

            if len(data[id_end_load]) != 0:
                for self.numberFrame in cd3:
                    data[id_end_load].pop()

            if len(data[id_start_unload]) != 0:
                for self.numberFrame in cd4:
                    data[id_start_unload].pop()

            if len(data[id_end_unload]) != 0:
                for self.numberFrame in cd5:
                    data[id_end_unload].pop()

        def saveSheetInDirectory(self,file):
            directorySave = askdirectory(title='Selecione onde deseja salvar o arquivo')

            if directorySave != '':
                file.save(directorySave + '/Horarios Motoristas.xlsx')
                self.end_word['text'] = 'Planilha Gerada!'
            else:
                self.end_word['text'] = 'Selecione uma pasta para salvar o arquivo'

        def creatSheet(self):
            cs1 = len(data[id_name])
            cs2 = len(data[id_start_load])
            cs3 = len(data[id_end_load])
            cs4 = len(data[id_start_unload])
            cs5 = len(data[id_end_unload])
 
            if cs1 and cs2 and cs3 and cs4 and cs5 != 0:
                file = Workbook()
                sheet = file.active
                sheet.title = 'OP TRANSPORTE'

                names_column = list(data.keys())

                for number_column, column in enumerate(names_column, start = 1):
                    cell = sheet.cell(row = 1, column = number_column)
                    cell.value = column

                    cell.fill = PatternFill(start_color = '0000CCFF', end_color = '0000CCFF', fill_type = 'solid')
                    cell.font = Font(bold = True, size = 13, name = 'Arial', color = '000000')
                    cell.alignment = Alignment(horizontal = 'center')

                    value_column = data[column]

                    for number_value, value in enumerate(value_column, start = 2):
                        cell = sheet.cell(row = number_value, column = number_column)
                        cell.value = value

                        cell.font = Font(name = 'Arial', size = 12)
                        cell.alignment = Alignment(horizontal = 'center')

                self.saveSheetInDirectory(file)
                self.clearDataBase()
            else:
                self.end_word['text'] = 'Insira as Informações'

        def insert(self):
            self.text_insert = 'Insira as Informações'

            if self.inputValue.get() != '':
                data[id_name].append(self.inputValue.get())
                self.inputValue.delete(0, 'end')
            else: 
                self.end_word['text'] = self.text_insert
                self.clearDataBase()

            if self.inputValue2.get() != '':
                data[id_start_load].append(self.inputValue2.get())
                self.inputValue2.delete(0, 'end')
            else: 
                self.end_word['text'] = self.text_insert
                self.clearDataBase()

            if self.inputValue3.get() != '':
                data[id_end_load].append(self.inputValue3.get())
                self.inputValue3.delete(0, 'end')
            else: 
                self.end_word['text'] = self.text_insert
                self.clearDataBase()

            if self.inputValue4.get() != '':
                data[id_start_unload].append(self.inputValue4.get())
                self.inputValue4.delete(0, 'end')
            else: 
                self.end_word['text'] = self.text_insert
                self.clearDataBase()

            if self.inputValue5.get() != '':
                data[id_end_unload].append(self.inputValue5.get())
                self.inputValue5.delete(0, 'end')
            else: 
                self.end_word['text'] = self.text_insert
                self.clearDataBase()
